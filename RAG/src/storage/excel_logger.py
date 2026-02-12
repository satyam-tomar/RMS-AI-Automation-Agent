# src/storage/excel_logger.py

from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook

from src.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelLogger:
    
    def __init__(self, excel_path: Path):
        self.excel_path = excel_path
        self._initialize_excel()
    
    def _initialize_excel(self) -> None:
        if not self.excel_path.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "Complaints"
            ws.append(["Timestamp", "Student Name", "Subject", "Complaint", "Resolution", "Status"])
            wb.save(str(self.excel_path))
            logger.info(f"✓ Created Excel: {self.excel_path}")
    
    def log_complaint(self, student_name: str, subject: str, complaint: str, resolution: str) -> bool:
        try:
            wb = load_workbook(str(self.excel_path))
            ws = wb.active
            
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([timestamp, student_name, subject, complaint, resolution, "Resolved"])
            
            wb.save(str(self.excel_path))
            logger.info(f"✓ Logged: {subject}")
            return True
        except Exception as e:
            logger.error(f"Excel error: {e}")
            return False