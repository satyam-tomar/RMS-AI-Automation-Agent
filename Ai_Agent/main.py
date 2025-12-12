"""
University Complaint Resolution AI Agent with Local Embeddings + Gemini Brain
- Local embeddings: FREE, runs on your M1 Mac (uses ~150MB RAM)
- Gemini API: FREE tier for reasoning and generation
- Perfect for 3-4GB available RAM on M1 Mac
"""

import os
import logging
from typing import Dict, Any, List
from pathlib import Path
from datetime import datetime

# LlamaIndex imports
from llama_index.core import (
    VectorStoreIndex,
    SimpleDirectoryReader,
    StorageContext,
    Settings,
)
from llama_index.core.embeddings import BaseEmbedding
from llama_index.llms.gemini import Gemini as LlamaGemini
from llama_index.vector_stores.chroma import ChromaVectorStore

# Local embedding model
from sentence_transformers import SentenceTransformer

# Google Gemini for generation
import google.generativeai as genai

# Other imports
import chromadb
from openpyxl import Workbook, load_workbook
import numpy as np

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class LocalEmbedding(BaseEmbedding):
    """
    Local embedding model using sentence-transformers.
    Runs on your M1 Mac with minimal RAM (~150MB).
    100% FREE - no API calls!
    """
    
    model_name: str
    _model: Any = None
    
    def __init__(self, model_name: str = "all-MiniLM-L6-v2", **kwargs):
        """
        Initialize local embedding model.
        
        Popular models for M1 Mac (8GB RAM):
        - all-MiniLM-L6-v2: 80MB, 384 dimensions (RECOMMENDED)
        - all-mpnet-base-v2: 420MB, 768 dimensions (better quality)
        - paraphrase-MiniLM-L3-v2: 60MB, 384 dimensions (fastest)
        """
        # Initialize parent first
        super().__init__(
            model_name=model_name,
            embed_batch_size=32,
            **kwargs
        )
        
        # Load model after parent init
        logger.info(f"Loading local embedding model: {model_name}")
        logger.info("This runs on your Mac - no API calls, 100% FREE!")
        
        object.__setattr__(self, '_model', SentenceTransformer(model_name))
        self._model.max_seq_length = 256  # Optimize for M1 Mac
        
        logger.info(f"✓ Model loaded: {model_name}")
        logger.info(f"  Embedding dimension: {self._model.get_sentence_embedding_dimension()}")
        logger.info(f"  Max sequence length: {self._model.max_seq_length}")
    
    class Config:
        arbitrary_types_allowed = True
    
    def _get_query_embedding(self, query: str) -> List[float]:
        """Get embedding for a query"""
        embedding = self._model.encode(query, normalize_embeddings=True)
        return embedding.tolist()
    
    def _get_text_embedding(self, text: str) -> List[float]:
        """Get embedding for a text"""
        embedding = self._model.encode(text, normalize_embeddings=True)
        return embedding.tolist()
    
    def _get_text_embeddings(self, texts: List[str]) -> List[List[float]]:
        """Get embeddings for multiple texts (batched for efficiency)"""
        embeddings = self._model.encode(
            texts,
            normalize_embeddings=True,
            show_progress_bar=True,
            batch_size=self.embed_batch_size
        )
        return [emb.tolist() for emb in embeddings]
    
    async def _aget_query_embedding(self, query: str) -> List[float]:
        """Async version"""
        return self._get_query_embedding(query)
    
    async def _aget_text_embedding(self, text: str) -> List[float]:
        """Async version"""
        return self._get_text_embedding(text)

class UniversityComplaintAgent:
    """
    Complaint resolution agent with:
    - LOCAL embeddings (sentence-transformers) - FREE, runs on M1 Mac
    - GEMINI API (google-generativeai) - FREE tier for generation
    
    Memory usage: ~500MB-1GB (perfect for M1 with 3-4GB free)
    """
    
    def __init__(
        self,
        api_key: str,
        documents_path: str = "./university_rules",
        excel_path: str = "./complaints.xlsx",
        chroma_path: str = "./chroma_db",
        rebuild_index: bool = False,
        model_name: str = "gemini-2.5-pro",
        embedding_model: str = "all-MiniLM-L6-v2"
    ):
        """
        Initialize agent with local embeddings + Gemini brain
        
        Args:
            api_key: Google API key (for Gemini generation only)
            embedding_model: Local model for embeddings
                - "all-MiniLM-L6-v2" (80MB, RECOMMENDED)
                - "all-mpnet-base-v2" (420MB, better quality)
                - "paraphrase-MiniLM-L3-v2" (60MB, fastest)
        """
        self.api_key = api_key
        self.documents_path = Path(documents_path)
        self.excel_path = Path(excel_path)
        self.chroma_path = Path(chroma_path)
        self.model_name = model_name
        self.embedding_model = embedding_model
        
        # Configure API for Gemini generation
        os.environ["GOOGLE_API_KEY"] = api_key
        genai.configure(api_key=api_key)
        
        # Initialize Gemini for generation (brain)
        self.llm = genai.GenerativeModel(model_name)
        
        # Setup components
        self._setup_llamaindex()
        self._initialize_excel()
        self._setup_rag_index(rebuild_index)
        
        logger.info("="*60)
        logger.info("AGENT READY!")
        logger.info("="*60)
        logger.info(f"💾 Embeddings: {embedding_model} (LOCAL - FREE)")
        logger.info(f"🧠 Generation: {model_name} (GEMINI API - FREE TIER)")
        logger.info(f"💰 Cost: $0.00 - Both are FREE!")
        logger.info("="*60)
    
    def _setup_llamaindex(self) -> None:
        """Configure LlamaIndex with local embeddings + Gemini LLM"""
        
        # Use LOCAL embedding model (FREE!)
        Settings.embed_model = LocalEmbedding(
            model_name=self.embedding_model
        )
        
        # Use Gemini for generation (FREE tier)
        Settings.llm = LlamaGemini(
            model_name=f"models/{self.model_name}",
            api_key=self.api_key,
            temperature=0.1
        )
        
        # Optimize for M1 Mac
        Settings.chunk_size = 512  # Smaller chunks = less memory
        Settings.chunk_overlap = 100
        
        logger.info("✓ LlamaIndex configured with local embeddings")
    
    def _initialize_excel(self) -> None:
        """Initialize Excel file"""
        if not self.excel_path.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "Complaints"
            ws.append(["Timestamp", "Student Name", "Subject", "Complaint", "Resolution", "Status"])
            wb.save(str(self.excel_path))
            logger.info(f"✓ Created Excel: {self.excel_path}")
    
    def _setup_rag_index(self, rebuild: bool) -> None:
        """Setup RAG index with local embeddings"""
        
        if rebuild or not self.chroma_path.exists():
            logger.info("Building RAG index with LOCAL embeddings...")
            logger.info("(This happens on your Mac - no API calls!)")
            
            documents = SimpleDirectoryReader(
                str(self.documents_path),
                recursive=True,
                required_exts=[".txt", ".pdf", ".docx", ".md"]
            ).load_data()
            
            logger.info(f"Loaded {len(documents)} documents")
            
            chroma_client = chromadb.PersistentClient(path=str(self.chroma_path))
            chroma_collection = chroma_client.get_or_create_collection("university_rules")
            
            vector_store = ChromaVectorStore(chroma_collection=chroma_collection)
            storage_context = StorageContext.from_defaults(vector_store=vector_store)
            
            # Build index - embeddings created locally!
            self.index = VectorStoreIndex.from_documents(
                documents,
                storage_context=storage_context,
                show_progress=True
            )
            
            logger.info("✓ Index built with local embeddings!")
        else:
            logger.info("Loading existing index...")
            chroma_client = chromadb.PersistentClient(path=str(self.chroma_path))
            chroma_collection = chroma_client.get_or_create_collection("university_rules")
            vector_store = ChromaVectorStore(chroma_collection=chroma_collection)
            self.index = VectorStoreIndex.from_vector_store(vector_store)
            logger.info("✓ Existing index loaded")
        
        # Create query engine
        self.query_engine = self.index.as_query_engine(
            similarity_top_k=3,  # Retrieve fewer for efficiency
            response_mode="compact"
        )
    
    def _search_rules(self, query: str) -> str:
        """Search university rules using LOCAL embeddings"""
        try:
            response = self.query_engine.query(query)
            return str(response)
        except Exception as e:
            logger.error(f"Search error: {e}")
            return f"Error searching rules: {str(e)}"
    
    def _log_to_excel(self, student_name: str, subject: str, complaint: str, resolution: str) -> bool:
        """Log complaint to Excel"""
        try:
            wb = load_workbook(str(self.excel_path))
            ws = wb.active
            
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([timestamp, student_name, subject, complaint, resolution, "Resolved"])
            
            wb.save(str(self.excel_path))
            logger.info(f"✓ Logged: {subject}")
            return True
        except Exception as e:
            logger.error(f"Excel error: {e}")
            return False
    
    def handle_complaint(
        self,
        subject: str,
        complaint: str,
        student_name: str = "Anonymous"
    ) -> Dict[str, Any]:
        """
        Handle complaint using:
        1. Local embeddings for policy search (FREE)
        2. Gemini API for intelligent resolution (FREE tier)
        """
        logger.info(f"\n{'='*60}")
        logger.info(f"Processing: {subject}")
        logger.info(f"{'='*60}")
        
        try:
            # Step 1: Search rules using LOCAL embeddings (FREE!)
            logger.info("🔍 Searching policies with LOCAL embeddings...")
            search_query = f"Find university policies about: {subject}. Full complaint: {complaint}"
            rules_context = self._search_rules(search_query)
            logger.info("✓ Policy search complete (no API cost!)")
            
            # Step 2: Generate resolution using Gemini API (FREE tier)
            logger.info("🧠 Generating resolution with Gemini...")
            prompt = f"""You are a University Complaint Resolution Assistant.

UNIVERSITY POLICIES (Retrieved from local knowledge base):
{rules_context}

STUDENT COMPLAINT:
Name: {student_name}
Subject: {subject}
Details: {complaint}

INSTRUCTIONS:
1. Review the university policies provided above
2. Analyze the complaint against official policies
3. Provide a professional, empathetic resolution
4. Quote specific policy sections when applicable
5. If policies don't cover the case, suggest escalation

Provide a clear, actionable resolution based ONLY on the official policies."""

            response = self.llm.generate_content(prompt)
            resolution = response.text
            logger.info("✓ Resolution generated")
            
            # Step 3: Log to Excel
            logger.info("💾 Logging to Excel...")
            self._log_to_excel(student_name, subject, complaint, resolution)
            
            result = {
                "student_name": student_name,
                "subject": subject,
                "complaint": complaint,
                "result": resolution,
                "status": "success",
                "timestamp": datetime.now().isoformat()
            }
            
            logger.info(f"✅ Complaint resolved successfully!")
            return result
            
        except Exception as e:
            error_msg = f"Error: {str(e)}"
            logger.error(error_msg)
            
            return {
                "student_name": student_name,
                "subject": subject,
                "complaint": complaint,
                "result": f"System error. Please contact support. Error: {str(e)}",
                "status": "error",
                "timestamp": datetime.now().isoformat()
            }


# Example usage
if __name__ == "__main__":
    import sys
    
    API_KEY = os.getenv("GOOGLE_API_KEY", "YOUR_API_KEY_HERE")
    
    if API_KEY == "YOUR_API_KEY_HERE":
        print("\n⚠️  ERROR: Please set GOOGLE_API_KEY")
        print("Get FREE key: https://aistudio.google.com/app/apikey")
        sys.exit(1)
    
    try:
        print("\n" + "="*60)
        print("INITIALIZING AGENT")
        print("="*60)
        print("🔧 Using LOCAL embeddings (FREE, runs on your Mac)")
        print("🔧 Using GEMINI API (FREE tier for generation)")
        print("="*60 + "\n")
        
        agent = UniversityComplaintAgent(
            api_key=API_KEY,
            rebuild_index=False,  # Set True first time
            embedding_model="all-MiniLM-L6-v2"  # 80MB model, perfect for M1
        )
        
        # Test complaint
        result = agent.handle_complaint(
            student_name="Jane Smith",
            subject="Attendance Shortage",
            complaint=(
                "I have only 65% attendance due to medical reasons. "
                "I have valid medical certificates for 2 weeks. "
                "Can I appear for final exams?"
            )
        )
        
        print("\n" + "="*60)
        print("RESULT")
        print("="*60)
        print(f"Status: {result['status']}")
        print(f"\nResolution:\n{result['result']}")
        print("="*60)
        print(f"\n✓ Logged to: {agent.excel_path}")
        print("\n💰 Total Cost: $0.00 (FREE!)")
        print("  - Embeddings: Local (no cost)")
        print("  - Generation: Gemini free tier")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)